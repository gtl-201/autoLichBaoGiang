import openpyxl



# START CREATE JOB
input_excel_job_path = './k10.xlsx'
plan_10 = openpyxl.load_workbook(input_excel_job_path)
plan_10_lst = plan_10.active
data_dict_10 = {}
for row in plan_10_lst.iter_rows(values_only=True):
    key = str(row[0])  # Assuming the key `is in column A (cột 1)
    data_dict_10[key] = row

input_excel_job_path = './k11.xlsx'
plan_11 = openpyxl.load_workbook(input_excel_job_path)
plan_11_lst = plan_11.active
data_dict_11 = {}
for row in plan_11_lst.iter_rows(values_only=True):
    key = str(row[0])  # Assuming the key `is in column A (cột 1)
    data_dict_11[key] = row

input_excel_job_path = './k12.xlsx'
plan_12 = openpyxl.load_workbook(input_excel_job_path)
plan_12_lst = plan_12.active
data_dict_12 = {}
for row in plan_12_lst.iter_rows(values_only=True):
    key = str(row[0])  # Assuming the key `is in column A (cột 1)
    data_dict_12[key] = row
# print(data_dict_11['31'][1])
# END CREATE JOB

input_excel_path = './mau.xlsx'
output_excel_path = './final.xlsx'

file = openpyxl.load_workbook(input_excel_path)

class_K10_lst = [] # Nhập các lớp dạy vào đây
class_K11_lst = ['11A1','11A2','11A3','11A4','11A5','11C1','11C2','11C3','11C4'] # Nhập các lớp dạy vào đây
class_k12_lst = ['12A2','12A3','12A4','12C2','12A1'] # Nhập các lớp dạy vào đây
count_K10_job = {class_val: 0 for class_val in class_K11_lst}
count_K11_job = {class_val: 0 for class_val in class_K11_lst}
count_K12_job = {class_val: 0 for class_val in class_k12_lst}

count_K10_job_tmp = {class_val: 0 for class_val in class_K11_lst}
count_K11_job_tmp = {class_val: 0 for class_val in class_K11_lst}
count_K12_job_tmp = {class_val: 0 for class_val in class_k12_lst}
tmp_TC_index = [4,5,6,10,11,12,16,17,18,22,23,24,28,29,30,34,35,36]

# print(count_job)
for sheet_name in file.sheetnames:
    sheet = file[sheet_name]
    # print(f'Dữ liệu ở cột E của sheet "{sheet_name}":')
    for row in range(1, sheet.max_row + 1):
        value_E = sheet.cell(row=row, column=5).value
        if value_E in class_K10_lst:
            count_K10_job[value_E] += 1
            if count_K10_job[value_E] in tmp_TC_index:
                sheet.cell(row=row, column=6).value = 'TC'
                sheet.cell(row=row, column=7).value = 'Luyen tap'
            else:
                count_K10_job_tmp[value_E] += 1
                sheet.cell(row=row, column=7).value = data_dict_10[f'{count_K10_job_tmp[value_E]}'][1]
                sheet.cell(row=row, column=6).value = count_K10_job_tmp[value_E]
        if value_E in class_K11_lst:
            count_K11_job[value_E] += 1
            if count_K11_job[value_E] in tmp_TC_index:
                sheet.cell(row=row, column=6).value = 'TC'
                sheet.cell(row=row, column=7).value = 'Luyen tap'
            else:
                count_K11_job_tmp[value_E] += 1
                sheet.cell(row=row, column=6).value = count_K11_job_tmp[value_E]
                sheet.cell(row=row, column=7).value = data_dict_11[f'{count_K11_job_tmp[value_E]}'][1]
        if value_E in class_k12_lst:
            count_K12_job[value_E] += 1
            if count_K12_job[value_E] in tmp_TC_index:
                sheet.cell(row=row, column=6).value = 'TC'
                sheet.cell(row=row, column=7).value = 'Luyen tap'
            else:
                count_K12_job_tmp[value_E] += 1
                sheet.cell(row=row, column=6).value = count_K12_job_tmp[value_E]
                sheet.cell(row=row, column=7).value = data_dict_12[f'{count_K12_job_tmp[value_E]}'][1]
            # print(f'Ô A{row}: {value_E}')


file.save(output_excel_path)